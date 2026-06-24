"""
Management command: sync_images
Reads the Excel workbook and updates image_url for internal places that
currently have no image, matching by place name.

Usage (inside the container):
    python manage.py sync_images
    python manage.py sync_images --dry-run
    python manage.py sync_images --excel /path/to/Lugares.xlsx
"""
import os
import re

from django.core.management.base import BaseCommand
from django.db.models import Q

from apps.places.models import Place

EXCEL_DEFAULT = r'c:\Users\lauta\OneDrive\Documentos\Lugares BS AS.xlsx'

FIELD_NUM = re.compile(r'^(\d+)[.\s]')

KEYWORD_FIELD = {
    'nombre del lugar': '1',
    'descripci': '2',
    'categor': '3',
    'tipo de actividad': '4',
    'imagen': '5',
    'ciudad': '6',
    'barrio': '7',
    'direcci': '8',
    'dias': '9',
    'horarios': '9',
    'presupuesto': '11',
    'gratis': '12',
    'reserva': '13',
    'sitio web': '14',
    'instagram': '14',
    'indoor': '15',
    'outdoor': '15',
    'sucursales': '16',
    'cantidad': '17',
    'lluvia': '18',
    'sol': '19',
    'ideal para': '20',
    'latitud': '21',
    'longitud': '22',
    'observaciones': '23',
}

# Domains that serve actual images (not web pages)
IMAGE_DOMAINS = (
    'googleusercontent.com',
    'encrypted-tbn0.gstatic.com',
    'tripadvisor.com/media',
    'wikimedia.org/wikipedia/commons',
    'feverup.com',
    'bairessecreta.com',
    'turismo.buenosaires.gob.ar/sites',
    'rocketcdn.me',
    'chicosxbaires.runasp.net',
    'agro.uba.ar',
    'upload.wikimedia.org',
    'offloadmedia',
)

IMAGE_EXTENSIONS = ('.jpg', '.jpeg', '.png', '.webp', '.gif')

DRIVE_FILE_RE = re.compile(r'drive\.google\.com/file/d/([a-zA-Z0-9_-]+)')
DRIVE_FOLDER_RE = re.compile(r'drive\.google\.com/drive/folders/')


def detect_field(cell_text):
    m = FIELD_NUM.match(cell_text)
    if m:
        num = m.group(1)
        colon = cell_text.find(':')
        val = cell_text[colon + 1:].strip() if colon != -1 else ''
        return num, val
    lower = cell_text.lower()
    colon = cell_text.find(':')
    if colon == -1:
        return None, None
    label = lower[:colon]
    val = cell_text[colon + 1:].strip()
    for kw, num in KEYWORD_FIELD.items():
        if kw in label:
            return num, val
    return None, None


def is_image_url(url: str) -> bool:
    if not url:
        return False
    url = url.strip()
    if not (url.startswith('http://') or url.startswith('https://')):
        return False
    # Google Drive folders → not usable
    if DRIVE_FOLDER_RE.search(url):
        return False
    # share.google links → Google Maps, not direct images
    if 'share.google/' in url:
        return False
    # Google Drive files → will be converted
    if DRIVE_FILE_RE.search(url):
        return True
    low = url.lower()
    # Direct image extension
    path = low.split('?')[0]
    if any(path.endswith(ext) for ext in IMAGE_EXTENSIONS):
        return True
    # Known image CDNs
    if any(d in low for d in IMAGE_DOMAINS):
        return True
    return False


def convert_image_url(url: str) -> str:
    """Convert Google Drive file links to direct-download URLs."""
    m = DRIVE_FILE_RE.search(url)
    if m:
        file_id = m.group(1)
        return f'https://drive.google.com/uc?export=view&id={file_id}'
    # Truncate to 500 chars to stay within URLField limits
    return url[:500]


def extract_places_from_excel(excel_path: str) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        raise ImportError('openpyxl is required: pip install openpyxl')

    wb = openpyxl.load_workbook(excel_path)
    sheets = [s for s in wb.sheetnames if s != 'EJEMPLO DE REFERENCIA']

    all_places = []

    for sheet_name in sheets:
        ws = wb[sheet_name]
        grid = {}
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    grid[(r, c)] = str(v).strip()

        barrio_row = None
        for r in range(1, 10):
            vals = [grid.get((r, c), '') for c in range(1, ws.max_column + 1)]
            non_empty = [v for v in vals if v]
            if len(non_empty) >= 3 and all(
                not FIELD_NUM.match(v)
                and 'DEBERAN' not in v.upper()
                and 'AGREGAR' not in v.upper()
                and len(v) < 50
                for v in non_empty
            ):
                barrio_row = r
                break
        if barrio_row is None:
            continue

        col_barrios = {
            c: grid.get((barrio_row, c), '')
            for c in range(1, ws.max_column + 1)
            if grid.get((barrio_row, c), '')
        }

        for col in col_barrios:
            col_cells = [
                grid.get((r, col), '')
                for r in range(barrio_row + 1, ws.max_row + 1)
                if grid.get((r, col), '')
            ]

            current_place = {}
            i = 0
            while i < len(col_cells):
                cell = col_cells[i]
                num, val = detect_field(cell)
                if num is not None:
                    if not val and i + 1 < len(col_cells):
                        next_cell = col_cells[i + 1]
                        next_num, _ = detect_field(next_cell)
                        if next_num is None:
                            val = next_cell
                            j = i + 2
                            while j < len(col_cells):
                                nn, _ = detect_field(col_cells[j])
                                if nn is not None:
                                    break
                                val += ' ' + col_cells[j]
                                j += 1
                            i = j - 1
                    if num == '1':
                        if current_place.get('1'):
                            all_places.append(current_place)
                        current_place = {'sheet': sheet_name}
                    if num not in current_place:
                        current_place[num] = val
                i += 1
            if current_place.get('1'):
                all_places.append(current_place)

    return all_places


def normalize_name(name: str) -> str:
    return re.sub(r'\s+', ' ', name.strip().lower())


class Command(BaseCommand):
    help = 'Sync image_url for internal places from the Excel workbook'

    def add_arguments(self, parser):
        parser.add_argument('--excel', default=EXCEL_DEFAULT, help='Path to the Excel file')
        parser.add_argument('--dry-run', action='store_true', help='Show what would be updated without saving')
        parser.add_argument('--overwrite', action='store_true', help='Overwrite existing image_url values too')

    def handle(self, *args, **options):
        excel_path = options['excel']
        dry_run = options['dry_run']
        overwrite = options['overwrite']

        if not os.path.exists(excel_path):
            self.stderr.write(self.style.ERROR(f'Excel not found: {excel_path}'))
            return

        self.stdout.write(f'Reading Excel: {excel_path}')
        places_raw = extract_places_from_excel(excel_path)
        self.stdout.write(f'Parsed {len(places_raw)} places from Excel')

        # Build name → image_url map (valid images only)
        excel_map: dict[str, str] = {}
        for p in places_raw:
            name = p.get('1', '').strip()
            raw_img = p.get('5', '').strip()
            if not name:
                continue
            if is_image_url(raw_img):
                img = convert_image_url(raw_img)
                key = normalize_name(name)
                # First occurrence wins (earlier sheets have priority)
                if key not in excel_map:
                    excel_map[key] = img

        self.stdout.write(f'Found {len(excel_map)} places with valid image URLs in Excel')

        # Match against DB
        db_places = Place.objects.filter(source='internal')
        if not overwrite:
            db_places = db_places.filter(Q(image_url='') | Q(image_url__isnull=True))

        updated = 0
        skipped = 0
        not_found = 0

        for place in db_places:
            key = normalize_name(place.name)
            if key in excel_map:
                new_url = excel_map[key]
                if dry_run:
                    self.stdout.write(f'  [DRY] {place.name[:45]} → {new_url[:80]}')
                else:
                    place.image_url = new_url
                    place.save(update_fields=['image_url'])
                    self.stdout.write(self.style.SUCCESS(f'  ✓ {place.name[:45]}'))
                updated += 1
            else:
                skipped += 1

        # Report places in Excel with images but not matched in DB
        db_names = {normalize_name(p.name) for p in Place.objects.filter(source='internal')}
        for name_key, img in excel_map.items():
            if name_key not in db_names:
                not_found += 1

        self.stdout.write('')
        self.stdout.write(f'Updated:   {updated}')
        self.stdout.write(f'No image:  {skipped} (places in DB without matching Excel image)')
        self.stdout.write(f'Not in DB: {not_found} (Excel places with images not found in DB)')
        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN — no changes saved'))
